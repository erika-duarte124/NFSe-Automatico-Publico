#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Leitor de NFS-e (padrão nacional SPED) — Gera relatório Excel
Lê todos os arquivos .xml de uma pasta e consolida as informações em planilha.

Retenções identificadas:
  - ISS   : vISSQN quando tpRetISSQN = 2  (tpRetISSQN=1 = NÃO retido)
  - INSS  : vRetCP
  - IRRF  : vRetIRRF
  - CSLL  : vRetCSLL
  - PIS   : vPis  (quando presente no tribFed/piscofins)
  - COFINS: vCofins (quando presente no tribFed/piscofins)
"""

import os
import sys
import glob
from xml.etree import ElementTree as ET
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─── NAMESPACE ──────────────────────────────────────────────────────────────
NS = "http://www.sped.fazenda.gov.br/nfse"

def ns(tag):
    return f"{{{NS}}}{tag}"

def find_text(element, path, default=""):
    """Navega no XML ignorando namespaces alternativos."""
    if element is None:
        return default
    # Tenta com namespace padrão
    parts = path.split("/")
    current = element
    for part in parts:
        found = current.find(ns(part))
        if found is None:
            # Tenta sem namespace (alguns XMLs omitem)
            found = current.find(part)
        if found is None:
            return default
        current = found
    return (current.text or "").strip() or default

def find_elem(element, path):
    """Retorna o elemento (não o texto)."""
    if element is None:
        return None
    parts = path.split("/")
    current = element
    for part in parts:
        found = current.find(ns(part))
        if found is None:
            found = current.find(part)
        if found is None:
            return None
        current = found
    return current

def to_float(value, default=0.0):
    try:
        return float(value.replace(",", "."))
    except (AttributeError, ValueError):
        return default

def fmt_cnpj(cnpj):
    """Formata CNPJ: 00.000.000/0000-00"""
    c = "".join(d for d in str(cnpj) if d.isdigit())
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
    return cnpj

def fmt_date(dt_str):
    """Converte datas ISO (2026-05-21T10:49:00-03:00) para DD/MM/AAAA."""
    if not dt_str:
        return ""
    try:
        return datetime.fromisoformat(dt_str[:10]).strftime("%d/%m/%Y")
    except ValueError:
        return dt_str[:10]

# ─── PARSER PRINCIPAL ────────────────────────────────────────────────────────
def parse_nfse(xml_path):
    """
    Lê um XML de NFS-e e retorna um dicionário com os campos do relatório.
    Retorna None se o arquivo não for uma NFS-e válida.
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"  [ERRO] Falha ao parsear {os.path.basename(xml_path)}: {e}")
        return None

    # Localiza infNFSe (pode estar na raiz ou aninhado)
    inf = find_elem(root, "infNFSe")
    if inf is None:
        # Tenta raiz direta
        if root.tag in (ns("infNFSe"), "infNFSe"):
            inf = root
        else:
            print(f"  [AVISO] infNFSe não encontrado em {os.path.basename(xml_path)}")
            return None

    # ── Dados do emissor (prestador) ────────────────────────────────────────
    emit       = find_elem(inf, "emit")
    cnpj_emit  = fmt_cnpj(find_text(emit, "CNPJ"))
    nome_emit  = find_text(emit, "xNome")

    # ── Número, chave e data ─────────────────────────────────────────────────
    n_nfse     = find_text(inf, "nNFSe")
    n_dfse     = find_text(inf, "nDFSe")   # número no sistema nacional
    # Chave está no atributo Id do infNFSe: "NFS" + 50 dígitos
    id_attr    = inf.get("Id", "")
    ch_nfse    = id_attr[3:] if id_attr.upper().startswith("NFS") else id_attr
    loc_incid  = find_text(inf, "xLocIncid")
    xTribNac   = find_text(inf, "xTribNac")

    # ── Valores do cabeçalho da nota (fonte mais confiável) ──────────────────
    vals_inf    = find_elem(inf, "valores")
    v_liq_hdr   = to_float(find_text(vals_inf, "vLiq"))
    v_bc_hdr    = to_float(find_text(vals_inf, "vBC"))
    v_issqn_hdr = to_float(find_text(vals_inf, "vISSQN"))

    # ── DPS (Declaração do Prestador de Serviços) ────────────────────────────
    dps  = find_elem(inf, "DPS")
    idps = find_elem(dps, "infDPS") if dps is not None else None

    # Data de emissão e competência
    dh_emi    = fmt_date(find_text(idps, "dhEmi"))
    d_compet  = fmt_date(find_text(idps, "dCompet"))

    # Tomador
    toma      = find_elem(idps, "toma")
    cnpj_toma = fmt_cnpj(find_text(toma, "CNPJ"))
    cpf_toma  = fmt_cnpj(find_text(toma, "CPF"))
    nome_toma = find_text(toma, "xNome")

    # Código de serviço (nacional) — item da lista de serviços
    c_trib_nac = find_text(idps, "serv/cServ/cTribNac")

    # Valor bruto do serviço
    vals_dps  = find_elem(idps, "valores")
    v_serv    = to_float(find_text(find_elem(vals_dps, "vServPrest"), "vServ"))
    if v_serv == 0.0:
        v_serv = v_bc_hdr  # fallback para notas municipais (SLIM HOTEL etc)

    # ── Tributos ─────────────────────────────────────────────────────────────
    trib      = find_elem(vals_dps, "trib")
    trib_mun  = find_elem(trib, "tribMun")
    trib_fed  = find_elem(trib, "tribFed")
    piscofins = find_elem(trib_fed, "piscofins")

    # ISS: só retido quando tpRetISSQN == "2"
    # Lê do cabeçalho (infNFSe/valores) que é o campo autoritativo;
    # cai para tribMun como fallback em notas que só preenchem o campo interno
    tp_ret_iss = (find_text(vals_inf, "tpRetISSQN")
                  or find_text(trib_mun, "tpRetISSQN"))
    p_aliq_iss = to_float(find_text(trib_mun, "pAliq"))

    if tp_ret_iss == "2":
        # Usa o valor do cabeçalho (mais confiável) ou calcula
        v_iss = v_issqn_hdr if v_issqn_hdr > 0 else round(v_serv * p_aliq_iss / 100, 2)
        iss_retido = True
    else:
        # ISS não retido pelo tomador; pode aparecer no cabeçalho como informativo
        v_iss = 0.0
        iss_retido = False

    # Alíquota ISS (informativa)
    aliq_iss_str = f"{p_aliq_iss:.2f}%".replace(".", ",") if p_aliq_iss > 0 else ""

    # INSS / CP
    v_ret_cp    = to_float(find_text(trib_fed, "vRetCP"))

    # IRRF
    v_ret_irrf  = to_float(find_text(trib_fed, "vRetIRRF"))

    # CSLL
    v_ret_csll  = to_float(find_text(trib_fed, "vRetCSLL"))

    # PIS / COFINS — só exibir quando efetivamente RETIDO.
    # tpRetPisCofins: "0" ou vazio = SEM retenção; "1".."9" = retido.
    # No XML, vPis/vCofins trazem a APURAÇÃO PRÓPRIA (aparecem mesmo sem
    # retenção), por isso só são levados para a planilha quando o código
    # indica retenção.
    tp_ret_pc    = find_text(piscofins, "tpRetPisCofins")
    cst_pc       = find_text(piscofins, "CST")
    houve_ret_pc = tp_ret_pc not in ("", "0")
    v_pis        = to_float(find_text(piscofins, "vPis"))    if houve_ret_pc else 0.0
    v_cofins     = to_float(find_text(piscofins, "vCofins")) if houve_ret_pc else 0.0

    # Total retenções: soma só o que foi efetivamente retido (inclui PIS/COFINS
    # apenas quando houve_ret_pc). NÃO usa v_total_ret_hdr (infNFSe/valores/
    # vTotalRet) porque esse campo pode conter o ISS informativo mesmo quando
    # tpRetISSQN=1 (não retido).
    v_total_ret = v_iss + v_ret_cp + v_ret_irrf + v_ret_csll + v_pis + v_cofins
    v_liq = v_liq_hdr if v_liq_hdr > 0 else (v_serv - v_total_ret)

    return {
        "arquivo":       os.path.basename(xml_path),
        "nNFSe":         n_nfse,
        "nDFSe":         n_dfse,
        "situacao":        "Normal",   # sobrescrito por gerar_retencoes com base nos eventos
        "codServ":      c_trib_nac,
        "chNFSe":        ch_nfse,
        "dhEmi":         dh_emi,
        "dCompet":       d_compet,
        "cnpjEmit":      cnpj_emit,
        "nomeEmit":      nome_emit,
        "cnpjToma":      cnpj_toma or cpf_toma,
        "nomeToma":      nome_toma,
        "xLocIncid":     loc_incid,
        "vServ":         v_serv,
        "aliqISS":       aliq_iss_str,
        "vISS":          v_iss,          # somente se retido (tpRetISSQN=2)
        "vINSS":         v_ret_cp,
        "vIRRF":         v_ret_irrf,
        "vCSLL":         v_ret_csll,
        "vPIS":          v_pis,
        "vCOFINS":       v_cofins,
        "cstPisCofins":  cst_pc,
        "tpRetPisCofins":tp_ret_pc,
        "vTotalRet":     v_total_ret,
        "vLiq":          v_liq,
    }

# ─── GERADOR DE EXCEL ────────────────────────────────────────────────────────
COLS = [
    ("Nº NFS-e",          "nNFSe",          14),
    ("Nº DFSe Nacional",  "nDFSe",          18),
    ("Status",            "situacao",         13),
    ("Data Emissão",      "dhEmi",          13),
    ("Competência",       "dCompet",        13),
    ("CNPJ Prestador",    "cnpjEmit",       20),
    ("Prestador",         "nomeEmit",       38),
    ("CNPJ Tomador",      "cnpjToma",       20),
    ("Tomador",           "nomeToma",       38),
    ("Município",         "xLocIncid",      16),
    ("Código Serviço",    "codServ",       14),
    ("Valor Bruto (R$)",  "vServ",          16),
    ("Alíq ISS",          "aliqISS",        10),
    ("ISS Retido (R$)",   "vISS",           15),
    ("INSS/CP (R$)",      "vINSS",          14),
    ("IRRF (R$)",         "vIRRF",          13),
    ("CSLL (R$)",         "vCSLL",          13),
    ("PIS (R$)",          "vPIS",           13),
    ("COFINS (R$)",       "vCOFINS",        14),
    ("CST PIS/COFINS",    "cstPisCofins",   14),
    ("Tipo Ret P/C",      "tpRetPisCofins", 13),
    ("Total Ret (R$)",    "vTotalRet",      15),
    ("Valor Líq (R$)",    "vLiq",           15),
    ("Chave NFS-e",       "chNFSe",         50),
]

MONEY_COLS = {"vServ","vISS","vINSS","vIRRF","vCSLL","vPIS","vCOFINS","vTotalRet","vLiq"}

# Cores
COR_CABECALHO  = PatternFill("solid", fgColor="1F497D")
COR_TOTAL      = PatternFill("solid", fgColor="D6E4F0")
COR_ZEBRA      = PatternFill("solid", fgColor="F2F7FB")
COR_CANCELADA  = PatternFill("solid", fgColor="FFC7CE")   # Status: Cancelada
COR_SUBST      = PatternFill("solid", fgColor="FFEB9C")   # Status: Substituída
FONTE_HDR      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FONTE_TOTAL    = Font(bold=True, size=10, name="Calibri")
FONTE_NORMAL   = Font(size=10, name="Calibri")

borda_fina = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

FMT_MOEDA = '#,##0.00'
FMT_PORCENTO = '0.00%'

def gerar_excel(registros, saida_path):
    wb = openpyxl.Workbook()

    # ── Aba de Dados ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "NFSe Recebidas"
    ws.freeze_panes = "A2"

    # Cabeçalho
    for col_idx, (header, _, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONTE_HDR
        cell.fill = COR_CABECALHO
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Dados
    for row_idx, reg in enumerate(registros, start=2):
        zebra = (row_idx % 2 == 0)
        for col_idx, (_, key, _) in enumerate(COLS, start=1):
            val = reg.get(key, "")
            # Converte 0.0 para vazio em campos monetários (menos vServ e vLiq)
            if key in MONEY_COLS and key not in ("vServ", "vLiq", "vTotalRet") and val == 0.0:
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONTE_NORMAL
            cell.border = borda_fina
            cell.alignment = Alignment(vertical="center")
            if key in MONEY_COLS:
                cell.number_format = FMT_MOEDA
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif key in ("nNFSe", "nDFSe", "situacao", "codServ"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if zebra and not key in MONEY_COLS:
                cell.fill = COR_ZEBRA
            # destaque do Status (prevalece sobre a zebra)
            if key == "situacao":
                if val == "Cancelada":
                    cell.fill = COR_CANCELADA
                elif val == "Substituída":
                    cell.fill = COR_SUBST

    # Linha de totais
    n_dados = len(registros)
    tot_row = n_dados + 2
    ws.cell(row=tot_row, column=1, value="TOTAL").font = FONTE_TOTAL

    soma_cols = ["vServ","vISS","vINSS","vIRRF","vCSLL","vPIS","vCOFINS","vTotalRet","vLiq"]
    for col_idx, (_, key, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=tot_row, column=col_idx)
        cell.border = borda_fina
        cell.fill = COR_TOTAL
        cell.font = FONTE_TOTAL
        if key in soma_cols:
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}2:{col_letter}{tot_row-1})"
            cell.value = formula
            cell.number_format = FMT_MOEDA
            cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ── Aba de Resumo ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    resumo_titulo = ws2.cell(row=1, column=1, value="Resumo de Retenções — NFSe")
    resumo_titulo.font = Font(bold=True, size=13, name="Calibri", color="1F497D")
    ws2.merge_cells("A1:B1")

    data_hora = ws2.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    data_hora.font = Font(italic=True, size=9, name="Calibri", color="7F7F7F")
    ws2.merge_cells("A2:B2")

    qtd = ws2.cell(row=3, column=1, value=f"Total de notas: {n_dados}")
    qtd.font = Font(size=10, name="Calibri")
    ws2.merge_cells("A3:B3")

    resumo_items = [
        ("Valor Bruto Total",    sum(r["vServ"]    for r in registros)),
        ("ISS Retido",           sum(r["vISS"]     for r in registros)),
        ("INSS / CP Retido",     sum(r["vINSS"]    for r in registros)),
        ("IRRF Retido",          sum(r["vIRRF"]    for r in registros)),
        ("CSLL Retido",          sum(r["vCSLL"]    for r in registros)),
        ("PIS (informativo)",    sum(r["vPIS"]     for r in registros)),
        ("COFINS (informativo)", sum(r["vCOFINS"]  for r in registros)),
        ("Total Retenções",      sum(r["vTotalRet"]for r in registros)),
        ("Valor Líquido Total",  sum(r["vLiq"]     for r in registros)),
    ]

    for i, (label, valor) in enumerate(resumo_items, start=5):
        separador = i in (5, 13)  # linha antes do total
        if i == 5 + len(resumo_items) - 2:  # linha antes de "Total Retenções"
            ws2.row_dimensions[i].height = 3

        cl = ws2.cell(row=i, column=1, value=label)
        cv = ws2.cell(row=i, column=2, value=valor)
        cv.number_format = FMT_MOEDA

        eh_total = label in ("Total Retenções", "Valor Líquido Total", "Valor Bruto Total")
        cl.font = Font(bold=eh_total, size=10, name="Calibri")
        cv.font = Font(bold=eh_total, size=10, name="Calibri")
        cv.alignment = Alignment(horizontal="right")

        if eh_total:
            cl.fill = COR_TOTAL
            cv.fill = COR_TOTAL

        cl.border = borda_fina
        cv.border = borda_fina

    wb.save(saida_path)
    print(f"\n✅  Excel gerado: {saida_path}")

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Leitor de NFS-e — Gerador de Relatório Excel")
    print("=" * 60)

    # Pasta com os XMLs
    while True:
        pasta = input("\nInforme o caminho da pasta com os arquivos XML:\n> ").strip().strip('"')
        if os.path.isdir(pasta):
            break
        print(f"  ❌ Pasta não encontrada: {pasta}")

    # Arquivo de saída
    default_saida = os.path.join(pasta, f"Relatorio_NFSe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    saida = input(f"\nCaminho do arquivo Excel de saída\n[Enter para salvar na mesma pasta: {default_saida}]\n> ").strip().strip('"')
    if not saida:
        saida = default_saida
    if not saida.lower().endswith(".xlsx"):
        saida += ".xlsx"

    # Lê os XMLs
    arquivos = sorted(glob.glob(os.path.join(pasta, "*.xml")))
    if not arquivos:
        print("\n❌ Nenhum arquivo .xml encontrado na pasta.")
        return

    print(f"\n📂 Encontrados {len(arquivos)} arquivo(s) XML. Processando...\n")
    registros = []
    erros = 0
    for arq in arquivos:
        print(f"  ▶ {os.path.basename(arq)}")
        reg = parse_nfse(arq)
        if reg:
            registros.append(reg)
        else:
            erros += 1

    if not registros:
        print("\n❌ Nenhuma NFS-e válida encontrada.")
        return

    # Ordena por data de emissão e número da nota
    def sort_key(r):
        try:
            d = datetime.strptime(r["dhEmi"], "%d/%m/%Y")
        except ValueError:
            d = datetime.min
        try:
            n = int(r["nNFSe"])
        except ValueError:
            n = 0
        return (d, n)

    registros.sort(key=sort_key)

    print(f"\n📊 {len(registros)} nota(s) processada(s) com sucesso. {erros} erro(s).")
    gerar_excel(registros, saida)

    # Resumo rápido no terminal
    print("\n── Resumo ─────────────────────────────────────────────")
    total_bruto = sum(r["vServ"]     for r in registros)
    total_ret   = sum(r["vTotalRet"] for r in registros)
    total_liq   = sum(r["vLiq"]      for r in registros)
    print(f"  Valor Bruto Total  : R$ {total_bruto:>12,.2f}")
    print(f"  Total Retenções    : R$ {total_ret:>12,.2f}")
    print(f"  Valor Líquido Total: R$ {total_liq:>12,.2f}")
    print("──────────────────────────────────────────────────────\n")

if __name__ == "__main__":
    main()
