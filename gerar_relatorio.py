# -*- coding: utf-8 -*-
"""
Relatório Simples de NFS-e em Excel, consultado direto do Portal Nacional.

Varre TODA a distribuição de documentos da empresa no ADN (desde o início,
sem depender dos arquivos já baixados) e gera uma planilha com duas abas:

  - Emitidas:  Número NFS-e, Data Geração, Competência, CNPJ/CPF Tomador,
               Nome Tomador, Município Emissor, Valor do Serviço, Situação,
               Chave NFS-e, XML Baixado, PDF Baixado
  - Recebidas: Número NFS-e, Data Geração, Competência, CNPJ/CPF Prestador,
               Nome Prestador, Valor do Serviço, Situação, Chave NFS-e,
               XML Baixado, PDF Baixado

Situação: Normal | Cancelada | Substituída (apurada pelos eventos da nota).
As colunas "XML/PDF Baixado" comparam o Portal com a pasta de saída — é a
conferência de que nada ficou faltando.

Uso:
  python gerar_relatorio.py --empresa "SGYM"
  python gerar_relatorio.py --empresa "SGYM" --competencia 2026-05
"""

import argparse
import getpass
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import baixar_nfse as nfse
import seguranca

# Eventos que mudam a situação da nota (códigos do Sistema Nacional NFS-e)
EVENTOS_CANCELAMENTO = {"101101", "101103"}
EVENTOS_SUBSTITUICAO = {"105102", "105104", "105105"}


# ------------------------------------------------------- extração de campos

def tag(xml: str, nome: str, bloco: str | None = None) -> str:
    """Retorna o conteúdo de uma tag; se 'bloco' for dado, procura só dentro dele."""
    trecho = xml
    if bloco:
        m = re.search(rf"<{bloco}>(.*?)</{bloco}>", xml, re.DOTALL)
        if not m:
            return ""
        trecho = m.group(1)
    m = re.search(rf"<{nome}>([^<]*)</{nome}>", trecho)
    return m.group(1).strip() if m else ""


def formatar_data(iso: str) -> str:
    """'2026-05-12T09:33:01-03:00' -> '12/05/2026'."""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", iso or "")
    return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else (iso or "")


def formatar_doc(numero: str) -> str:
    """Formata CNPJ (14) ou CPF (11) com a pontuação usual."""
    n = re.sub(r"\D", "", numero or "")
    if len(n) == 14:
        return f"{n[:2]}.{n[2:5]}.{n[5:8]}/{n[8:12]}-{n[12:]}"
    if len(n) == 11:
        return f"{n[:3]}.{n[3:6]}.{n[6:9]}-{n[9:]}"
    return numero or ""


def extrair_nota(xml: str, chave: str) -> dict:
    """Extrai os campos do Relatório Simples do XML de uma NFS-e."""
    emit_doc = tag(xml, "CNPJ", "emit") or tag(xml, "CPF", "emit")
    emit_nome = tag(xml, "xNome", "emit")
    toma_doc = tag(xml, "CNPJ", "toma") or tag(xml, "CPF", "toma")
    toma_nome = tag(xml, "xNome", "toma")
    valor = tag(xml, "vServ") or tag(xml, "vLiq")
    try:
        valor = float(valor)
    except (TypeError, ValueError):
        valor = None
    # mês de EMISSÃO (dhEmi); se faltar, cai para o mês embutido na chave
    dh_emi = tag(xml, "dhEmi")
    m_emi = re.match(r"(\d{4}-\d{2})", dh_emi or "")
    emissao = m_emi.group(1) if m_emi else nfse.info_da_chave(chave)["competencia"]
    return {
        "chave": chave,
        "numero": (tag(xml, "nNFSe") or nfse.info_da_chave(chave)["numero"]).lstrip("0") or "0",
        "data_geracao": formatar_data(tag(xml, "dhProc") or tag(xml, "dhEmi")),
        "competencia": tag(xml, "dCompet"),
        "emissao": emissao,
        "municipio_emissor": tag(xml, "xLocEmi"),
        "emit_doc": emit_doc,
        "emit_nome": emit_nome,
        "toma_doc": toma_doc,
        "toma_nome": toma_nome,
        "valor": valor,
        "situacao": "Normal",
    }


def codigo_evento(xml: str) -> str:
    """Descobre o código do evento (ex.: 101101) dentro do XML de um evento."""
    m = re.search(r"<e(\d{6})\b", xml) or re.search(r"<(?:tpEvento|cEvento)>(\d{6})", xml)
    return m.group(1) if m else ""


def chave_do_evento(xml: str) -> str:
    m = re.search(r"<chNFSe>(\d{50})</chNFSe>", xml)
    return m.group(1) if m else ""


# ------------------------------------------------------------ coleta no ADN

def coletar_do_portal(empresa: dict, competencia: str | None) -> tuple[dict, dict]:
    """
    Percorre toda a distribuição da empresa no ADN (NSU 0 em diante) sem gravar
    arquivos. Retorna (notas por chave, situações por chave).
    """
    nome = empresa["nome"]
    senha = seguranca.revelar(empresa.get("senha")) or getpass.getpass(f"Senha do certificado de {nome}: ")
    sessao = nfse.criar_sessao(empresa["certificado"], senha)

    notas: dict[str, dict] = {}
    situacoes: dict[str, str] = {}
    nsu = 0
    confirmacoes_fim = 0
    while True:
        url = nfse.URL_ADN_DISTRIBUICAO.format(nsu=nsu)
        resp = nfse.obter_com_retentativas(sessao, url)
        if resp is None:
            raise RuntimeError(
                f"Portal indisponível durante a varredura (NSU {nsu}); "
                f"relatório NÃO gerado para não sair incompleto.")
        if resp.status_code == 404:
            confirmacoes_fim += 1
            if confirmacoes_fim < 3:
                time.sleep(10)
                continue
            break
        confirmacoes_fim = 0
        if resp.status_code != 200:
            nfse.log(f"  HTTP {resp.status_code} (NSU {nsu}): {nfse.descrever_erro(resp)}")
            break

        docs = nfse.processar_lote(resp.json())
        if not docs:
            break

        for doc in docs:
            if not doc["xml_b64"]:
                continue
            try:
                xml = nfse.descompactar_xml(doc["xml_b64"]).decode("utf-8", errors="replace")
            except Exception:
                continue

            if doc["tipo"].upper().startswith("EVENTO"):
                cod = codigo_evento(xml)
                ch = chave_do_evento(xml) or doc["chave"]
                if cod in EVENTOS_CANCELAMENTO:
                    situacoes[ch] = "Cancelada"
                elif cod in EVENTOS_SUBSTITUICAO:
                    situacoes[ch] = "Substituída"
            else:
                nota = extrair_nota(xml, doc["chave"])
                # filtra pelo mês de EMISSÃO (dhEmi): mostra tudo que foi emitido
                # no mês, inclusive notas retroativas (competência de mês anterior)
                if competencia and (nota["emissao"] or "")[:7] != competencia:
                    continue
                notas[doc["chave"]] = nota

        nfse.log(f"  NSU {nsu} -> {max(d['nsu'] for d in docs)}: "
                 f"{len(notas)} nota(s) e {len(situacoes)} evento(s) de situação até agora")
        maior = max(d["nsu"] for d in docs)
        if maior <= nsu:
            break
        nsu = maior
        # varredura mais lenta que o download de propósito: rajadas de consulta
        # rápidas são o que dispara o limitador de tráfego do servidor (503)
        time.sleep(4.0)

    for ch, situacao in situacoes.items():
        if ch in notas:
            notas[ch]["situacao"] = situacao
    return notas, situacoes


# ------------------------------------------------------------------- planilha

CABECALHO_EMITIDAS = ["Número NFS-e", "Data Geração", "Competência", "CNPJ/CPF Tomador",
                      "Nome Tomador", "Município Emissor", "Valor do Serviço",
                      "Situação", "Chave NFS-e", "XML Baixado", "PDF Baixado"]
CABECALHO_RECEBIDAS = ["Número NFS-e", "Data Geração", "Competência", "CNPJ/CPF Prestador",
                       "Nome Prestador", "Valor do Serviço",
                       "Situação", "Chave NFS-e", "XML Baixado", "PDF Baixado"]

AZUL = PatternFill("solid", fgColor="1F4E78")
VERMELHO_CLARO = PatternFill("solid", fgColor="FFC7CE")
AMARELO_CLARO = PatternFill("solid", fgColor="FFEB9C")
AMARELO_RETRO = PatternFill("solid", fgColor="FFF2CC")   # linha inteira: nota retroativa
CINZA_TOTAL = PatternFill("solid", fgColor="D9E1F2")


def preencher_aba(aba, cabecalho, linhas, retroativas=None):
    aba.append(cabecalho)
    for cel in aba[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = AZUL
        cel.alignment = Alignment(horizontal="center")
    col_valor = cabecalho.index("Valor do Serviço") + 1
    col_situacao = cabecalho.index("Situação") + 1
    primeira = 2                       # primeira linha de dados
    for idx, linha in enumerate(linhas):
        aba.append(linha)
        r = aba.max_row
        aba.cell(row=r, column=col_valor).number_format = "#,##0.00"
        # nota retroativa (competência de mês anterior à emissão): linha em amarelo
        if retroativas and retroativas[idx]:
            for c in range(1, len(cabecalho) + 1):
                aba.cell(row=r, column=c).fill = AMARELO_RETRO
        # a cor da situação prevalece sobre o amarelo da retroativa
        situacao = aba.cell(row=r, column=col_situacao)
        if situacao.value == "Cancelada":
            situacao.fill = VERMELHO_CLARO
        elif situacao.value == "Substituída":
            situacao.fill = AMARELO_CLARO
    ultima = aba.max_row               # última linha de dados (antes do TOTAL)
    larguras = {"Número NFS-e": 14, "Data Geração": 13, "Competência": 13,
                "CNPJ/CPF Tomador": 20, "Nome Tomador": 40, "CNPJ/CPF Prestador": 20,
                "Nome Prestador": 40, "Município Emissor": 24, "Valor do Serviço": 16,
                "Situação": 13, "Chave NFS-e": 54, "XML Baixado": 13, "PDF Baixado": 13}
    for i, titulo in enumerate(cabecalho, start=1):
        aba.column_dimensions[get_column_letter(i)].width = larguras.get(titulo, 16)
    aba.freeze_panes = "A2"

    ncols = len(cabecalho)
    if linhas:
        # o filtro cobre só cabeçalho + dados; a linha de TOTAL fica de fora
        aba.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ultima}"
        # TOTAL do faturamento com SUBTOTAL(9,...): soma só o que estiver
        # VISÍVEL, então acompanha os filtros que a Erika aplicar na coluna
        letra = get_column_letter(col_valor)
        total_row = ultima + 1
        rot = aba.cell(row=total_row, column=1, value="TOTAL")
        rot.font = Font(bold=True)
        rot.fill = CINZA_TOTAL
        cel_total = aba.cell(row=total_row, column=col_valor,
                             value=f"=SUBTOTAL(9,{letra}{primeira}:{letra}{ultima})")
        cel_total.number_format = "#,##0.00"
        cel_total.font = Font(bold=True)
        cel_total.fill = CINZA_TOTAL
    else:
        aba.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def gerar_planilha(empresa: dict, pasta_saida: Path, notas: dict,
                   competencia: str | None) -> Path:
    cnpj_empresa = nfse.somente_digitos(empresa["cnpj"]).lstrip("0")
    pasta_empresa = pasta_saida / empresa["nome"]

    xml_baixados = {p.stem for p in pasta_empresa.rglob("*.xml")} if pasta_empresa.exists() else set()
    pdf_baixados = {p.stem for p in pasta_empresa.rglob("*.pdf")} if pasta_empresa.exists() else set()

    emitidas, recebidas = [], []
    emitidas_retro, recebidas_retro = [], []
    for n in sorted(notas.values(), key=lambda x: (x["competencia"], x["numero"].zfill(15))):
        baixado_xml = "Sim" if n["chave"] in xml_baixados else "NÃO"
        baixado_pdf = "Sim" if n["chave"] in pdf_baixados else "NÃO"
        eh_emitida = nfse.somente_digitos(n["emit_doc"]).lstrip("0") == cnpj_empresa
        # retroativa: competência (dCompet) de mês diferente do mês de emissão
        comp_mes = (n["competencia"] or "")[:7]
        retroativa = bool(comp_mes) and comp_mes != (n["emissao"] or "")[:7]
        if eh_emitida:
            emitidas.append([n["numero"], n["data_geracao"], n["competencia"],
                             formatar_doc(n["toma_doc"]), n["toma_nome"],
                             n["municipio_emissor"], n["valor"], n["situacao"],
                             n["chave"], baixado_xml, baixado_pdf])
            emitidas_retro.append(retroativa)
        else:
            recebidas.append([n["numero"], n["data_geracao"], n["competencia"],
                              formatar_doc(n["emit_doc"]), n["emit_nome"],
                              n["valor"], n["situacao"],
                              n["chave"], baixado_xml, baixado_pdf])
            recebidas_retro.append(retroativa)

    wb = Workbook()
    preencher_aba(wb.active, CABECALHO_EMITIDAS, emitidas, emitidas_retro)
    wb.active.title = "Emitidas"
    preencher_aba(wb.create_sheet("Recebidas"), CABECALHO_RECEBIDAS, recebidas, recebidas_retro)

    sufixo = f"_{competencia}" if competencia else ""
    pasta_relatorios = (pasta_empresa / f"{competencia}_Relatorios" if competencia
                        else pasta_empresa / "Relatorios")
    destino = pasta_relatorios / f"Relatorio_Simples_{empresa['nome']}{sufixo}_{datetime.now():%Y-%m-%d_%H%M}.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)

    nfse.log(f"  Emitidas: {len(emitidas)} | Recebidas: {len(recebidas)}")
    faltam_xml = sum(1 for lista in (emitidas, recebidas) for l in lista if l[-2] == "NÃO")
    faltam_pdf = sum(1 for lista in (emitidas, recebidas) for l in lista if l[-1] == "NÃO")
    if faltam_xml or faltam_pdf:
        nfse.log(f"  ATENÇÃO: {faltam_xml} XML(s) e {faltam_pdf} PDF(s) constam no "
                 f"Portal mas não estão na pasta (veja as colunas 'Baixado').")
    else:
        nfse.log("  Conferência OK: todas as notas do Portal estão baixadas na pasta.")
    nfse.log(f"  Planilha salva em: {destino}")
    return destino


# ----------------------------------------------------------------- principal

def main() -> int:
    parser = argparse.ArgumentParser(description="Gera o Relatório Simples (Excel) das NFS-e direto do Portal.")
    parser.add_argument("--empresa", help="filtra empresas pelo nome (contém)")
    parser.add_argument("--competencia", help="só notas do mês informado (ex.: 2026-05 ou 05/2026)")
    args = parser.parse_args()

    config = nfse.carregar_json(nfse.ARQ_CONFIG, {})
    if not config:
        print(f"Crie o arquivo de configuração: {nfse.ARQ_CONFIG}")
        return 1
    nfse.definir_ambiente(config.get("ambiente", "producao"))
    pasta_saida = Path(config.get("pasta_saida", nfse.PASTA_SCRIPT / "notas"))

    empresas = config.get("empresas", [])
    if args.empresa:
        empresas = [e for e in empresas if args.empresa.lower() in e["nome"].lower()]
    if not empresas:
        print("Nenhuma empresa encontrada na configuração com esse filtro.")
        return 1

    competencia = nfse.normalizar_competencia(args.competencia) if args.competencia else None

    falhas = 0
    for empresa in empresas:
        nfse.log(f"==== Relatório Simples: {empresa['nome']} ====")
        try:
            notas, _ = coletar_do_portal(empresa, competencia)
            gerar_planilha(empresa, pasta_saida, notas, competencia)
        except Exception as exc:
            nfse.log(f"  ERRO em {empresa['nome']}: {exc}")
            falhas += 1
    return 1 if falhas else 0


if __name__ == "__main__":
    sys.exit(main())
