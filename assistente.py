# -*- coding: utf-8 -*-
"""
Assistente de configuração — tela inicial que qualquer pessoa usa para
cadastrar suas empresas, sem editar nenhum arquivo de código. Ao final,
gera o config.json que os outros scripts (rotina.py, rodar_fila.py etc.)
já sabem ler.

Etapa 2 do projeto público: pasta de destino, cadastro de empresas (com
validação de certificado e aviso de vencimento), período inicial e
frequência de execução. Ao concluir, grava o config.json completo que
rotina.py e rodar_fila.py já sabem ler.
"""

import json
import re
import shutil
import subprocess
import sys
import tkinter as tk
import webbrowser
from datetime import date
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from cryptography.hazmat.primitives.serialization import pkcs12
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import despacho
import seguranca

SUBCOMANDOS = {
    "baixar_nfse", "gerar_relatorio", "gerar_relatorio_pdf",
    "gerar_retencoes", "rotina", "rodar_fila", "backfill", "executar_agora",
}

PASTA = despacho.PASTA
ARQ_CONFIG = PASTA / "config.json"
PASTA_SAIDA_PADRAO = str(Path.home() / "Downloads" / "NFSe-Automatico")
LINKEDIN_URL = "https://www.linkedin.com/in/erika-duarte-tech/"
LIMITE_EMPRESAS = 20
LIMITE_GRUPOS = 3
LIMITE_POR_GRUPO = 10
GRUPO_PADRAO = "Grupo 1"
GRUPOS_FIXOS = [f"Grupo {i}" for i in range(1, LIMITE_GRUPOS + 1)]

AGENDAMENTO_PADRAO = {
    "mensal":    {"ativo": True,  "dia_mes": 1, "hora": "09:00"},
    "semanal":   {"ativo": False, "dia_semana": 2, "hora": "09:00"},
    "quinzenal": {"ativo": False, "dia_semana": 2, "hora": "09:00"},
}

DIAS_SEMANA = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
               "Sexta-feira", "Sábado", "Domingo"]
HORAS = [f"{h:02d}:00" for h in range(24)]
MESES = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]


def limpar_cnpj(texto: str) -> str:
    return re.sub(r"\D", "", texto)


def validar_certificado(caminho: str, senha: str) -> tuple[bool, str, str | None]:
    """Tenta abrir o .pfx localmente com a senha informada, sem acessar a
    internet. Retorna (ok, mensagem, validade AAAA-MM-DD ou None)."""
    arq = Path(caminho)
    if not arq.exists():
        return False, "Arquivo não encontrado.", None
    try:
        dados = arq.read_bytes()
        _, cert, _ = pkcs12.load_key_and_certificates(dados, senha.encode("utf-8") if senha else None)
        validade = cert.not_valid_after_utc.date().isoformat() if cert is not None else None
        return True, "Certificado válido.", validade
    except ValueError:
        return False, "Senha incorreta ou arquivo de certificado inválido.", None
    except Exception as e:
        return False, f"Erro ao ler o certificado: {e}", None


def certificados_vencidos(empresas: list[dict]) -> list[str]:
    hoje = date.today().isoformat()
    return [e["nome"] for e in empresas if e.get("cert_validade") and e["cert_validade"] < hoje]


NOME_TAREFA_PREFIXO = "NFSe Automatico"


def registrar_tarefas_agendador(agendamentos_por_grupo: dict) -> list[str]:
    """Cria (ou remove, se desativada) 1 tarefa por grupo+frequência no
    Agendador de Tarefas do Windows, via PowerShell Register-ScheduledTask —
    tarefa diária, com StartWhenAvailable (recupera atraso se o PC estava
    desligado). Retorna lista de mensagens de erro (vazia se tudo OK)."""
    erros = []

    for grupo, agendamento in agendamentos_por_grupo.items():
        for freq, cfg in agendamento.items():
            nome_tarefa = f"{NOME_TAREFA_PREFIXO} - {grupo} - {freq.capitalize()}"
            if not cfg.get("ativo"):
                subprocess.run(
                    ["powershell", "-NoProfile", "-Command",
                     f'Unregister-ScheduledTask -TaskName "{nome_tarefa}" -Confirm:$false -ErrorAction SilentlyContinue'],
                    capture_output=True, text=True)
                continue

            comando = despacho.comando_base(silencioso=True) + ["rotina", "--grupo", grupo, "--modo", freq]
            argumentos = " ".join(f'"{a}"' if " " in a else a for a in comando[1:])
            script = (
                f"$acao = New-ScheduledTaskAction -Execute '{comando[0]}' "
                f"-Argument '{argumentos}' -WorkingDirectory '{PASTA}'; "
                f"$gatilho = New-ScheduledTaskTrigger -Daily -At {cfg['hora']}; "
                f"$config = New-ScheduledTaskSettingsSet -StartWhenAvailable -DontStopIfGoingOnBatteries; "
                f'Register-ScheduledTask -TaskName "{nome_tarefa}" -Action $acao -Trigger $gatilho '
                f"-Settings $config -Force | Out-Null"
            )
            r = subprocess.run(["powershell", "-NoProfile", "-Command", script],
                               capture_output=True, text=True)
            if r.returncode != 0:
                erros.append(f"{nome_tarefa}: {r.stderr.strip() or r.stdout.strip()}")

    return erros


class Assistente(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NFS-e Automático — configuração inicial")
        self.geometry("640x600")
        self.resizable(False, False)

        self.pasta_saida = tk.StringVar(value=PASTA_SAIDA_PADRAO)
        self.empresas: list[dict] = []          # cada empresa tem uma chave "grupo"
        self.agendamentos_por_grupo: dict = {}  # nome do grupo -> agendamento
        self.editando_index: int | None = None
        self.avisar_cert_vencido = True
        self.periodo_inicial = {"tipo": "completo"}

        self.eh_primeira_execucao = not ARQ_CONFIG.exists()
        if not self.eh_primeira_execucao:
            config_existente = json.loads(ARQ_CONFIG.read_text(encoding="utf-8"))
            self.pasta_saida.set(config_existente.get("pasta_saida", PASTA_SAIDA_PADRAO).replace("/", "\\"))
            self.avisar_cert_vencido = config_existente.get("avisar_cert_vencido", True)
            self.periodo_inicial = config_existente.get("periodo_inicial", self.periodo_inicial)
            for grupo in config_existente.get("grupos", []):
                nome_grupo = grupo.get("nome", GRUPO_PADRAO)
                self.agendamentos_por_grupo[nome_grupo] = grupo.get("agendamento", AGENDAMENTO_PADRAO)
                for emp in grupo.get("empresas", []):
                    emp = dict(emp)
                    emp["grupo"] = nome_grupo
                    self.empresas.append(emp)
        self.nomes_empresas_iniciais = {e["nome"] for e in self.empresas}
        # grupos que já tinham pelo menos 1 frequência ATIVA salva no disco antes
        # desta sessão — usado pra avisar antes de sobrescrever na Tela 4 (um
        # grupo só com o cadastro salvo parcialmente, sem nunca ter passado
        # pelo Concluir, não conta — não existe tarefa real pra sobrescrever)
        self.grupos_com_agendamento_previo = {
            g for g, ag in self.agendamentos_por_grupo.items()
            if any(v.get("ativo") for v in ag.values())
        }
        self._aviso_agendamento_mostrado = False
        self._montar_rodape_contato()

        self.container = tk.Frame(self)
        self.container.pack(fill="both", expand=True)

        if self.eh_primeira_execucao:
            self.mostrar_tela_pasta()
        else:
            self.mostrar_tela_empresas()
        self.after(200, self.verificar_certificados_vencidos)

    def _montar_rodape_contato(self):
        rodape = tk.Frame(self)
        rodape.pack(side="bottom", fill="x", pady=(0, 6))
        centro = tk.Frame(rodape)
        centro.pack(anchor="center")
        tk.Label(centro, text="Desenvolvido por Erika Duarte  —  ", fg="#777", font=("Segoe UI", 8)).pack(side="left")
        link = tk.Label(centro, text="LinkedIn", fg="#0a66c2", font=("Segoe UI", 8, "underline"), cursor="hand2")
        link.pack(side="left")
        link.bind("<Button-1>", lambda e: webbrowser.open(LINKEDIN_URL))

    def verificar_certificados_vencidos(self):
        if not self.avisar_cert_vencido:
            return
        vencidos = certificados_vencidos(self.empresas)
        if not vencidos:
            return

        janela = tk.Toplevel(self)
        janela.title("Certificados vencidos")
        janela.resizable(False, False)
        janela.grab_set()

        tk.Label(janela, text="Certificados vencidos:", font=("Segoe UI", 10, "bold")).pack(padx=20, pady=(16, 4), anchor="w")
        for nome in vencidos:
            tk.Label(janela, text=f"  •  {nome}").pack(padx=20, anchor="w")
        tk.Label(janela, text="Caso não queira receber esse aviso, retire a empresa\n"
                              "do cadastro de empresas (ou renove o certificado).",
                 justify="left", fg="#555").pack(padx=20, pady=(10, 12), anchor="w")

        var_nao_avisar = tk.BooleanVar(value=False)
        tk.Checkbutton(janela, text="Não mostrar este aviso novamente",
                       variable=var_nao_avisar).pack(padx=20, anchor="w")

        def fechar():
            if var_nao_avisar.get():
                self.avisar_cert_vencido = False
                self._salvar_preferencia_aviso()
            janela.destroy()

        tk.Button(janela, text="OK", width=10, command=fechar).pack(pady=16)
        janela.transient(self)

    def _salvar_preferencia_aviso(self):
        if not ARQ_CONFIG.exists():
            return
        config_existente = json.loads(ARQ_CONFIG.read_text(encoding="utf-8"))
        config_existente["avisar_cert_vencido"] = False
        ARQ_CONFIG.write_text(json.dumps(config_existente, indent=2, ensure_ascii=False), encoding="utf-8")

    def _mostrar_limite_atingido(self):
        janela = tk.Toplevel(self)
        janela.title("Limite de empresas")
        janela.resizable(False, False)
        janela.grab_set()

        tk.Label(janela, text="Esse é um programa gratuito :)", font=("Segoe UI", 11, "bold")).pack(padx=24, pady=(20, 6))
        tk.Label(janela, justify="center", wraplength=340, text=
                 f"Ele foi feito pra ajudar escritórios de pequeno e médio porte — por isso o "
                 f"cadastro vai até {LIMITE_EMPRESAS} empresas.\n"
                 "Caso precise de algum suporte maior, entre em contato com a desenvolvedora."
                 ).pack(padx=24, pady=(0, 14))

        centro = tk.Frame(janela)
        centro.pack(pady=(0, 4))
        tk.Label(centro, text="Link LinkedIn:  ").pack(side="left")
        link = tk.Label(centro, text="LinkedIn", fg="#0a66c2", font=("Segoe UI", 10, "underline"), cursor="hand2")
        link.pack(side="left")
        link.bind("<Button-1>", lambda e: webbrowser.open(LINKEDIN_URL))

        tk.Label(janela, text="Developed by: Erika Duarte", font=("Segoe UI", 9)).pack(pady=(2, 0))

        tk.Button(janela, text="Entendi", width=12, command=janela.destroy).pack(pady=(10, 20))
        janela.transient(self)

    def limpar_container(self):
        for w in self.container.winfo_children():
            w.destroy()

    # ---------------------------------------------------------- Tela 1
    def mostrar_tela_pasta(self):
        self.limpar_container()
        f = self.container
        tk.Label(f, text="Onde guardar as notas baixadas?", font=("Segoe UI", 13, "bold")).pack(pady=(24, 8))
        tk.Label(f, text="Todo o histórico de NFS-e (XML, PDF e relatórios) será salvo\n"
                          "dentro dessa pasta, organizado por empresa e por mês.",
                 justify="center").pack(pady=(0, 16))

        linha = tk.Frame(f)
        linha.pack(pady=8)
        entrada = tk.Entry(linha, textvariable=self.pasta_saida, width=55)
        entrada.pack(side="left", padx=(0, 8))
        tk.Button(linha, text="Escolher pasta...", command=self.escolher_pasta).pack(side="left")

        tk.Button(f, text="Avançar  →", font=("Segoe UI", 10, "bold"),
                  command=self.mostrar_tela_empresas).pack(pady=32)

    def escolher_pasta(self):
        escolhida = filedialog.askdirectory(title="Escolher pasta para salvar as notas")
        if escolhida:
            self.pasta_saida.set(str(Path(escolhida) / "NFSe-Automatico"))

    # ---------------------------------------------------------- Tela 2
    def mostrar_tela_empresas(self):
        Path(self.pasta_saida.get()).mkdir(parents=True, exist_ok=True)

        self.limpar_container()
        f = self.container
        tk.Label(f, text="Cadastro de empresas", font=("Segoe UI", 13, "bold")).pack(pady=(16, 4))

        form = tk.LabelFrame(f, text="Nova empresa", padx=12, pady=10)
        form.pack(fill="x", padx=16, pady=8)

        tk.Label(form, text="Nome da empresa").grid(row=0, column=0, sticky="w")
        self.var_nome = tk.StringVar()
        tk.Entry(form, textvariable=self.var_nome, width=45).grid(row=0, column=1, columnspan=2, sticky="w")

        tk.Label(form, text="CNPJ").grid(row=1, column=0, sticky="w", pady=(6, 0))
        self.var_cnpj = tk.StringVar()
        tk.Entry(form, textvariable=self.var_cnpj, width=25).grid(row=1, column=1, sticky="w", pady=(6, 0))

        tk.Label(form, text="Grupo").grid(row=2, column=0, sticky="w", pady=(6, 0))
        self.var_grupo = tk.StringVar(value=self._grupos_existentes()[0] if self._grupos_existentes() else GRUPO_PADRAO)
        self.combo_grupo = ttk.Combobox(form, textvariable=self.var_grupo, width=22,
                                         values=GRUPOS_FIXOS, state="readonly")
        self.combo_grupo.grid(row=2, column=1, sticky="w", pady=(6, 0))
        tk.Label(form, text="(até 3 grupos, 10 empresas cada)",
                 fg="#777", font=("Segoe UI", 8)).grid(row=2, column=2, sticky="w", padx=(8, 0), pady=(6, 0))

        tk.Label(form, text="Certificado (.pfx)").grid(row=3, column=0, sticky="w", pady=(6, 0))
        self.var_cert_label = tk.StringVar(value="(nenhum arquivo escolhido)")
        self.var_cert_caminho = ""
        tk.Label(form, textvariable=self.var_cert_label, fg="#555", anchor="w", width=32).grid(row=3, column=1, sticky="w", pady=(6, 0))
        tk.Button(form, text="Escolher...", command=self.escolher_certificado).grid(row=3, column=2, sticky="w", pady=(6, 0))

        tk.Label(form, text="Senha do certificado").grid(row=4, column=0, sticky="w", pady=(6, 0))
        self.var_senha = tk.StringVar()
        tk.Entry(form, textvariable=self.var_senha, show="*", width=25).grid(row=4, column=1, sticky="w", pady=(6, 0))

        self.var_status = tk.StringVar(value="")
        tk.Label(form, textvariable=self.var_status, anchor="w").grid(row=5, column=0, columnspan=3, sticky="w", pady=(8, 0))

        botoes = tk.Frame(form)
        botoes.grid(row=6, column=0, columnspan=3, sticky="w", pady=(10, 0))
        tk.Button(botoes, text="Validar certificado", command=self.acao_validar).pack(side="left")
        self.btn_adicionar = tk.Button(botoes, text="Adicionar empresa", state="disabled", command=self.acao_adicionar)
        self.btn_adicionar.pack(side="left", padx=(8, 0))

        self.lista_frame = tk.LabelFrame(f, text="Empresas cadastradas", padx=8, pady=8)
        lista_frame = self.lista_frame
        lista_frame.pack(fill="both", expand=True, padx=16, pady=8)

        self.tree = ttk.Treeview(lista_frame, columns=("nome", "cnpj", "grupo", "status"), show="headings", height=7)
        self.tree.heading("nome", text="Empresa")
        self.tree.heading("cnpj", text="CNPJ")
        self.tree.heading("grupo", text="Grupo")
        self.tree.heading("status", text="Certificado")
        self.tree.column("nome", width=210)
        self.tree.column("cnpj", width=130)
        self.tree.column("grupo", width=80)
        self.tree.column("status", width=90)
        self.tree.pack(side="left", fill="both", expand=True)

        botoes_lista = tk.Frame(lista_frame)
        botoes_lista.pack(side="left", padx=(8, 0), fill="y")
        tk.Button(botoes_lista, text="Editar", width=10, command=self.acao_editar).pack(pady=2)
        tk.Button(botoes_lista, text="Remover", width=10, command=self.acao_remover).pack(pady=2)

        rodape = tk.Frame(f)
        rodape.pack(pady=12, fill="x", padx=16)
        tk.Button(rodape, text="←  Voltar", command=self.mostrar_tela_pasta).pack(side="left", padx=(0, 8))
        self.btn_concluir = tk.Button(rodape, text="Avançar  →", font=("Segoe UI", 10, "bold"),
                                       state="disabled", command=self.mostrar_tela_periodo)
        self.btn_concluir.pack(side="left")
        tk.Button(rodape, text="Histórico de execuções...", command=self.acao_ver_historico).pack(side="right")
        tk.Button(rodape, text="Rodar agora (mês específico)...", command=self.acao_rodar_agora).pack(side="right", padx=(0, 8))

        self._atualizar_lista()

    def escolher_certificado(self):
        caminho = filedialog.askopenfilename(title="Escolher certificado digital",
                                              filetypes=[("Certificado A1", "*.pfx *.p12"), ("Todos os arquivos", "*.*")])
        if caminho:
            self.var_cert_caminho = caminho
            self.var_cert_label.set(Path(caminho).name)
            self.var_status.set("")

    def acao_validar(self) -> bool:
        if not self.var_cert_caminho:
            self.var_status.set("⚠ Escolha um arquivo de certificado primeiro.")
            return False
        ok, msg, validade = validar_certificado(self.var_cert_caminho, self.var_senha.get())
        if ok and validade:
            msg += f" (válido até {date.fromisoformat(validade):%d/%m/%Y})"
        self.var_status.set(("✓ " if ok else "✗ ") + msg)
        self.var_status_ok = ok
        self.var_cert_validade = validade
        self.btn_adicionar.config(state="normal" if ok else "disabled")
        return ok

    def _grupos_existentes(self) -> list[str]:
        vistos = []
        for emp in self.empresas:
            g = emp.get("grupo", GRUPO_PADRAO)
            if g not in vistos:
                vistos.append(g)
        return vistos

    def _config_atual(self) -> dict:
        grupos_config = []
        for grupo in self._grupos_existentes():
            empresas_do_grupo = []
            for e in self.empresas:
                if e.get("grupo", GRUPO_PADRAO) == grupo:
                    e2 = dict(e)
                    e2.pop("grupo", None)
                    empresas_do_grupo.append(e2)
            grupos_config.append({
                "nome": grupo,
                # {} pra grupo que ainda não passou pela Tela 4 — não inventa um
                # agendamento "ativo" no disco só porque a empresa foi salva cedo
                "agendamento": self.agendamentos_por_grupo.get(grupo, {}),
                "empresas": empresas_do_grupo,
            })
        return {
            "ambiente": "producao",
            "pasta_saida": self.pasta_saida.get().replace("\\", "/"),
            "avisar_cert_vencido": self.avisar_cert_vencido,
            "periodo_inicial": self.periodo_inicial,
            "grupos": grupos_config,
        }

    def _salvar_config(self) -> None:
        """Grava o config.json com o estado atual — chamado a cada empresa
        adicionada/removida (não só no Concluir), pra não perder o cadastro
        se a pessoa fechar o programa no meio do caminho."""
        ARQ_CONFIG.write_text(json.dumps(self._config_atual(), indent=2, ensure_ascii=False), encoding="utf-8")

    def acao_adicionar(self):
        nome = self.var_nome.get().strip()
        cnpj = limpar_cnpj(self.var_cnpj.get())
        grupo = self.var_grupo.get().strip() or GRUPO_PADRAO
        if not nome or not cnpj:
            messagebox.showwarning("Faltam dados", "Preencha o nome e o CNPJ da empresa.")
            return
        if len(cnpj) != 14:
            messagebox.showwarning("CNPJ inválido", "O CNPJ deve ter 14 dígitos. Confira o número digitado.")
            return
        if not getattr(self, "var_status_ok", False):
            messagebox.showwarning("Certificado não validado", "Clique em \"Validar certificado\" antes de adicionar.")
            return

        empresas_sem_esta = [e for i, e in enumerate(self.empresas) if i != self.editando_index]

        duplicada = next((e for e in empresas_sem_esta if e["cnpj"] == cnpj), None)
        if duplicada:
            messagebox.showwarning("CNPJ já cadastrado",
                                    f"O CNPJ {cnpj} já está cadastrado como \"{duplicada['nome']}\" "
                                    f"(Grupo: {duplicada.get('grupo', GRUPO_PADRAO)}).")
            return

        if len(empresas_sem_esta) >= LIMITE_EMPRESAS:
            self._mostrar_limite_atingido()
            return

        grupos_existentes = {e.get("grupo", GRUPO_PADRAO) for e in empresas_sem_esta}
        if grupo not in grupos_existentes and len(grupos_existentes) >= LIMITE_GRUPOS:
            messagebox.showwarning("Limite de grupos",
                                    f"Este programa aceita no máximo {LIMITE_GRUPOS} grupos. "
                                    f"Grupos já usados: {', '.join(sorted(grupos_existentes))}.")
            return

        no_grupo = sum(1 for e in empresas_sem_esta if e.get("grupo", GRUPO_PADRAO) == grupo)
        if no_grupo >= LIMITE_POR_GRUPO:
            messagebox.showwarning("Limite do grupo",
                                    f"O grupo \"{grupo}\" já tem {no_grupo} empresas (máximo {LIMITE_POR_GRUPO}). "
                                    "Escolha outro grupo (ou crie um novo, se ainda não atingiu o limite de "
                                    f"{LIMITE_GRUPOS} grupos).")
            return

        pasta_certs = Path(self.pasta_saida.get()) / "Certificados"
        pasta_certs.mkdir(parents=True, exist_ok=True)
        destino = pasta_certs / f"{cnpj}.pfx"
        shutil.copy2(self.var_cert_caminho, destino)

        registro = {
            "nome": nome,
            "cnpj": cnpj,
            "grupo": grupo,
            "certificado": str(destino).replace("\\", "/"),
            "senha": seguranca.proteger(self.var_senha.get()),
            "usar_cnpj_consulta": False,
            "cert_validade": getattr(self, "var_cert_validade", None),
        }
        if self.editando_index is not None:
            self.empresas[self.editando_index] = registro
            self.editando_index = None
        else:
            self.empresas.append(registro)

        self._limpar_formulario()
        self._atualizar_lista()
        self._salvar_config()

    def acao_editar(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        emp = self.empresas[idx]
        self.var_nome.set(emp["nome"])
        self.var_cnpj.set(emp["cnpj"])
        self.var_grupo.set(emp.get("grupo", GRUPO_PADRAO))
        self.var_cert_caminho = emp["certificado"]
        self.var_cert_label.set(Path(emp["certificado"]).name)
        self.var_senha.set(seguranca.revelar(emp["senha"]))
        self.var_status.set("")
        self.btn_adicionar.config(state="disabled")
        self.var_status_ok = False
        self.editando_index = idx

    def acao_remover(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        if messagebox.askyesno("Remover", f"Remover \"{self.empresas[idx]['nome']}\" da lista?"):
            del self.empresas[idx]
            self._atualizar_lista()
            self._salvar_config()

    def acao_rodar_agora(self):
        """Abre o diálogo de retirada manual de um mês específico — roda por
        fora do agendamento, não atrasa nem antecipa os fechamentos automáticos."""
        if not self.empresas:
            messagebox.showinfo("Rodar agora", "Cadastre pelo menos uma empresa primeiro.")
            return

        janela = tk.Toplevel(self)
        janela.title("Rodar agora — mês específico")
        janela.resizable(False, False)
        janela.grab_set()

        tk.Label(janela, text="Retirada manual de um mês específico",
                 font=("Segoe UI", 11, "bold")).pack(padx=20, pady=(18, 4))
        tk.Label(janela, justify="center", fg="#777", text=
                 "Roda por fora do agendamento — não atrasa nem antecipa\n"
                 "as retiradas automáticas. Útil pra pegar uma nota atrasada\n"
                 "de um mês que já foi fechado."
                 ).pack(padx=20, pady=(0, 14))

        form = tk.Frame(janela)
        form.pack(padx=20)

        tk.Label(form, text="Empresa:").grid(row=0, column=0, sticky="w", pady=4)
        nomes = [e["nome"] for e in self.empresas]
        var_empresa = tk.StringVar(value=nomes[0])
        ttk.Combobox(form, textvariable=var_empresa, values=nomes, state="readonly", width=30).grid(
            row=0, column=1, sticky="w", pady=4)

        tk.Label(form, text="Mês:").grid(row=1, column=0, sticky="w", pady=4)
        linha_data = tk.Frame(form)
        linha_data.grid(row=1, column=1, sticky="w", pady=4)
        hoje = date.today()
        var_mes = tk.StringVar(value=f"{hoje.month:02d}")
        var_ano = tk.StringVar(value=str(hoje.year))
        ttk.Combobox(linha_data, textvariable=var_mes, values=MESES, width=5, state="readonly").pack(side="left")
        anos = [str(a) for a in range(hoje.year - 4, hoje.year + 1)]
        ttk.Combobox(linha_data, textvariable=var_ano, values=anos, width=7, state="readonly").pack(
            side="left", padx=(4, 0))

        def executar():
            empresa = var_empresa.get()
            competencia = f"{var_ano.get()}-{var_mes.get()}"
            janela.destroy()
            subprocess.Popen(despacho.comando_base() + ["executar_agora", "--empresa", empresa,
                                                          "--competencia", competencia], cwd=str(PASTA))
            messagebox.showinfo("Rodando em segundo plano",
                                 f"Buscando notas de \"{empresa}\" — competência {competencia}.\n\n"
                                 "Isso roda em segundo plano e não trava a tela; um aviso do Windows\n"
                                 "avisa quando terminar.")

        tk.Button(janela, text="Executar", command=executar).pack(pady=(6, 18))
        janela.transient(self)

    def acao_ver_historico(self):
        """Mostra o histórico das últimas execuções (manuais e automáticas),
        lido de ultima_execucao.json, com opção de exportar pra Excel."""
        arq_execucao = PASTA / "ultima_execucao.json"
        dados = json.loads(arq_execucao.read_text(encoding="utf-8")) if arq_execucao.exists() else {}
        grupos = dados.get("grupos", {})

        janela = tk.Toplevel(self)
        janela.title("Histórico de execuções")
        janela.geometry("680x420")
        janela.grab_set()

        tk.Label(janela, text="Histórico de execuções", font=("Segoe UI", 12, "bold")).pack(pady=(14, 4))

        if not grupos:
            tk.Label(janela, text="Nenhuma execução automática registrada ainda.", fg="#777").pack(pady=30)
            tk.Button(janela, text="Fechar", command=janela.destroy).pack(pady=10)
            janela.transient(self)
            return

        frame_tree = tk.Frame(janela)
        frame_tree.pack(fill="both", expand=True, padx=16, pady=8)
        colunas = [("grupo", "Grupo", 70), ("rotulo", "Execução", 140), ("competencia", "Competência", 90),
                   ("data", "Data", 90), ("empresa", "Empresa", 190), ("status", "Status", 70)]
        tree = ttk.Treeview(frame_tree, columns=[c[0] for c in colunas], show="headings", height=13)
        for chave, titulo, largura in colunas:
            tree.heading(chave, text=titulo)
            tree.column(chave, width=largura)
        tree.pack(side="left", fill="both", expand=True)
        scroll = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side="left", fill="y")
        tree.tag_configure("falha", background="#FFC7CE")

        linhas = []
        for grupo, info in grupos.items():
            rotulo = info.get("rotulo", "-")
            competencia = info.get("competencia", "-")
            data_exec = info.get("data", "-")
            for nome in info.get("empresas_ok", []):
                linha = (grupo, rotulo, competencia, data_exec, nome, "OK")
                linhas.append(linha)
                tree.insert("", "end", values=linha)
            for nome in info.get("empresas_com_falha", []):
                linha = (grupo, rotulo, competencia, data_exec, nome, "FALHOU")
                linhas.append(linha)
                tree.insert("", "end", values=linha, tags=("falha",))

        def exportar():
            destino = filedialog.asksaveasfilename(
                title="Salvar histórico em Excel", defaultextension=".xlsx",
                filetypes=[("Planilha Excel", "*.xlsx")],
                initialfile=f"Historico_Execucoes_{date.today():%Y-%m-%d}.xlsx")
            if not destino:
                return
            wb = Workbook()
            aba = wb.active
            aba.title = "Histórico"
            aba.append(["Grupo", "Execução", "Competência", "Data", "Empresa", "Status"])
            for cel in aba[1]:
                cel.font = Font(bold=True, color="FFFFFF")
                cel.fill = PatternFill("solid", fgColor="1F4E78")
            for linha in linhas:
                aba.append(list(linha))
                if linha[-1] == "FALHOU":
                    for c in range(1, 7):
                        aba.cell(row=aba.max_row, column=c).fill = PatternFill("solid", fgColor="FFC7CE")
            for i, largura in enumerate([12, 20, 14, 12, 34, 10], start=1):
                aba.column_dimensions[get_column_letter(i)].width = largura
            aba.freeze_panes = "A2"
            wb.save(destino)
            messagebox.showinfo("Exportado", f"Histórico salvo em:\n{destino}")

        rodape_hist = tk.Frame(janela)
        rodape_hist.pack(pady=10)
        tk.Button(rodape_hist, text="Exportar para Excel...", command=exportar).pack(side="left", padx=(0, 8))
        tk.Button(rodape_hist, text="Fechar", command=janela.destroy).pack(side="left")
        janela.transient(self)

    def _limpar_formulario(self):
        self.var_nome.set("")
        self.var_cnpj.set("")
        self.var_senha.set("")
        self.var_cert_caminho = ""
        self.var_cert_label.set("(nenhum arquivo escolhido)")
        self.var_status.set("")
        self.var_status_ok = False
        self.var_cert_validade = None
        self.btn_adicionar.config(state="disabled")

    def _atualizar_lista(self):
        self.tree.delete(*self.tree.get_children())
        for emp in self.empresas:
            self.tree.insert("", "end", values=(emp["nome"], emp["cnpj"], emp.get("grupo", GRUPO_PADRAO), "✓ Válido"))
        self.btn_concluir.config(state="normal" if self.empresas else "disabled")
        self.lista_frame.config(text=f"Empresas cadastradas ({len(self.empresas)}/{LIMITE_EMPRESAS})")
        if not self.var_grupo.get():
            self.var_grupo.set(GRUPO_PADRAO)

    # ---------------------------------------------------------- Tela 3
    def mostrar_tela_periodo(self):
        self.limpar_container()
        f = self.container
        tk.Label(f, text="A partir de quando buscar as notas?", font=("Segoe UI", 13, "bold")).pack(pady=(24, 8))
        tk.Label(f, text="Essa escolha vale para a primeira busca (o histórico).\n"
                          "Depois disso, o programa passa a buscar só o que for novo.",
                 justify="center").pack(pady=(0, 20))

        self.var_periodo_tipo = tk.StringVar(value=self.periodo_inicial.get("tipo", "completo"))

        tk.Radiobutton(f, text="Buscar todo o histórico da empresa", variable=self.var_periodo_tipo,
                       value="completo", command=self._atualizar_campos_periodo).pack(anchor="w", padx=60, pady=4)
        tk.Radiobutton(f, text="Buscar a partir de um mês específico:", variable=self.var_periodo_tipo,
                       value="mes_especifico", command=self._atualizar_campos_periodo).pack(anchor="w", padx=60, pady=4)

        linha = tk.Frame(f)
        linha.pack(anchor="w", padx=90, pady=(0, 20))
        hoje = date.today()
        desde = self.periodo_inicial.get("desde", f"{hoje.year}-{hoje.month:02d}")
        ano_desde, mes_desde = desde.split("-")
        self.var_mes = tk.StringVar(value=mes_desde)
        self.var_ano = tk.StringVar(value=ano_desde)
        anos = [str(a) for a in range(hoje.year - 4, hoje.year + 1)]
        self.combo_mes = ttk.Combobox(linha, textvariable=self.var_mes, values=MESES, width=5, state="readonly")
        self.combo_mes.pack(side="left", padx=(0, 6))
        self.combo_ano = ttk.Combobox(linha, textvariable=self.var_ano, values=anos, width=7, state="readonly")
        self.combo_ano.pack(side="left")

        rodape = tk.Frame(f)
        rodape.pack(pady=20)
        tk.Button(rodape, text="←  Voltar", command=self.mostrar_tela_empresas).pack(side="left", padx=(0, 8))
        tk.Button(rodape, text="Avançar  →", font=("Segoe UI", 10, "bold"),
                  command=self._salvar_periodo_e_avancar).pack(side="left")

        self._atualizar_campos_periodo()

    def _atualizar_campos_periodo(self):
        estado = "readonly" if self.var_periodo_tipo.get() == "mes_especifico" else "disabled"
        self.combo_mes.config(state=estado)
        self.combo_ano.config(state=estado)

    def _salvar_periodo_e_avancar(self):
        if self.var_periodo_tipo.get() == "completo":
            self.periodo_inicial = {"tipo": "completo"}
        else:
            self.periodo_inicial = {"tipo": "mes_especifico", "desde": f"{self.var_ano.get()}-{self.var_mes.get()}"}
        self.mostrar_tela_frequencia()

    # ---------------------------------------------------------- Tela 4
    def mostrar_tela_frequencia(self):
        self.limpar_container()
        f = self.container
        tk.Label(f, text="Com que frequência cada grupo roda?", font=("Segoe UI", 13, "bold")).pack(pady=(14, 4))
        tk.Label(f, text="Cada grupo tem sua própria agenda — até 2 frequências ativas por grupo.",
                 justify="center").pack(pady=(0, 8))

        grupos = self._grupos_existentes()
        notebook = ttk.Notebook(f)
        notebook.pack(fill="both", expand=True, padx=16, pady=4)

        self.vars_freq_por_grupo = {}
        self.vars_freq_detalhe_por_grupo = {}

        for grupo in grupos:
            aba = tk.Frame(notebook)
            notebook.add(aba, text=grupo)
            self._montar_aba_frequencia(aba, grupo)

        grupos_ja_agendados = [g for g in grupos if g in self.grupos_com_agendamento_previo]
        if grupos_ja_agendados and not self._aviso_agendamento_mostrado:
            self._aviso_agendamento_mostrado = True
            messagebox.showinfo("Agendamento já existente",
                                 f"Já existe agendamento configurado para: {', '.join(grupos_ja_agendados)}.\n\n"
                                 "As configurações abaixo serão substituídas pelas novas ao clicar em Concluir.")

        tk.Label(f, text="Dica: prefira horários fora do expediente comercial (fim de tarde, noite ou\n"
                          "manhã cedo) — a API do governo costuma ficar mais instável durante o dia.",
                 fg="#777", font=("Segoe UI", 8), justify="center").pack(pady=(6, 0))

        rodape = tk.Frame(f)
        rodape.pack(pady=14)
        tk.Button(rodape, text="←  Voltar", command=self.mostrar_tela_periodo).pack(side="left", padx=(0, 8))
        tk.Button(rodape, text="Concluir  →", font=("Segoe UI", 10, "bold"),
                  command=self.acao_finalizar).pack(side="left")

    def _montar_aba_frequencia(self, aba, grupo):
        # "or" (não .get com default) porque um grupo salvo cedo (antes de
        # passar pela Tela 4) tem entrada {} no dicionário — {} é "falsy",
        # então cai no AGENDAMENTO_PADRAO do mesmo jeito que se não existisse
        agendamento = self.agendamentos_por_grupo.get(grupo) or AGENDAMENTO_PADRAO
        vars_freq = {}
        vars_detalhe = {}

        def bloco(chave, titulo, tem_dia_mes):
            cfg = agendamento.get(chave, {})
            grp = tk.LabelFrame(aba, text=titulo, padx=12, pady=8)
            grp.pack(fill="x", padx=14, pady=6)

            var_ativo = tk.BooleanVar(value=cfg.get("ativo", False))
            vars_freq[chave] = var_ativo
            tk.Checkbutton(grp, text="Ativar", variable=var_ativo,
                           command=lambda: self._limitar_frequencias(grupo, chave)).grid(row=0, column=0, sticky="w")

            if tem_dia_mes:
                tk.Label(grp, text="Dia do mês:").grid(row=0, column=1, sticky="e", padx=(16, 4))
                var_dia = tk.StringVar(value=str(cfg.get("dia_mes", 1)))
                ttk.Combobox(grp, textvariable=var_dia, values=[str(d) for d in range(1, 29)],
                            width=4, state="readonly").grid(row=0, column=2)
                detalhe = {"dia_mes": var_dia}
            else:
                tk.Label(grp, text="Dia da semana:").grid(row=0, column=1, sticky="e", padx=(16, 4))
                var_dia = tk.StringVar(value=DIAS_SEMANA[cfg.get("dia_semana", 2)])
                ttk.Combobox(grp, textvariable=var_dia, values=DIAS_SEMANA,
                            width=13, state="readonly").grid(row=0, column=2)
                detalhe = {"dia_semana": var_dia}

            tk.Label(grp, text="Horário:").grid(row=0, column=3, sticky="e", padx=(16, 4))
            var_hora = tk.StringVar(value=cfg.get("hora", "09:00"))
            ttk.Combobox(grp, textvariable=var_hora, values=HORAS, width=6, state="readonly").grid(row=0, column=4)
            detalhe["hora"] = var_hora
            vars_detalhe[chave] = detalhe

        bloco("mensal", "Mensal — fecha o mês anterior completo", tem_dia_mes=True)
        bloco("semanal", "Semanal — toda semana, mês atual", tem_dia_mes=False)
        bloco("quinzenal", "Quinzenal — a cada 14 dias, mês atual", tem_dia_mes=False)

        n_empresas = sum(1 for e in self.empresas if e.get("grupo", GRUPO_PADRAO) == grupo)
        tk.Label(aba, text=f"{n_empresas} empresa(s) neste grupo.",
                 fg="#777", font=("Segoe UI", 8)).pack(pady=(4, 0))

        self.vars_freq_por_grupo[grupo] = vars_freq
        self.vars_freq_detalhe_por_grupo[grupo] = vars_detalhe

    def _limitar_frequencias(self, grupo, chave_alterada):
        vars_freq = self.vars_freq_por_grupo[grupo]
        ativas = [k for k, v in vars_freq.items() if v.get()]
        if len(ativas) > 2:
            vars_freq[chave_alterada].set(False)
            messagebox.showwarning("Limite de frequências", "Escolha no máximo 2 frequências ativas por grupo.")

    def acao_finalizar(self):
        agendamentos_por_grupo = {}
        for grupo, vars_freq in self.vars_freq_por_grupo.items():
            if not any(v.get() for v in vars_freq.values()):
                messagebox.showwarning("Nenhuma frequência ativa",
                                        f'Ative pelo menos 1 frequência para o grupo "{grupo}" antes de concluir.')
                return
            agendamento = {}
            for chave, var_ativo in vars_freq.items():
                detalhe = self.vars_freq_detalhe_por_grupo[grupo][chave]
                item = {"ativo": var_ativo.get(), "hora": detalhe["hora"].get()}
                if "dia_mes" in detalhe:
                    item["dia_mes"] = int(detalhe["dia_mes"].get())
                else:
                    item["dia_semana"] = DIAS_SEMANA.index(detalhe["dia_semana"].get())
                agendamento[chave] = item
            agendamentos_por_grupo[grupo] = agendamento
        self.agendamentos_por_grupo = agendamentos_por_grupo
        self._salvar_config()

        self.config(cursor="watch")
        self.update()
        erros = registrar_tarefas_agendador(self.agendamentos_por_grupo)
        nomes_novas = [e["nome"] for e in self.empresas if e["nome"] not in self.nomes_empresas_iniciais]
        if nomes_novas:
            subprocess.Popen(despacho.comando_base() + ["backfill", "--empresas", ",".join(nomes_novas)], cwd=str(PASTA))
        self.config(cursor="")

        total_empresas = len(self.empresas)
        resumo_grupos = ", ".join(
            f'{g} ({sum(1 for e in self.empresas if e.get("grupo", GRUPO_PADRAO) == g)})'
            for g in self._grupos_existentes())
        resumo_periodo = ("todo o histórico" if self.periodo_inicial.get("tipo") == "completo"
                          else f"a partir de {self.periodo_inicial.get('desde')}")
        if erros:
            messagebox.showwarning("Configuração concluída (com avisos)",
                                    f"{total_empresas} empresa(s) em {len(self._grupos_existentes())} grupo(s): {resumo_grupos}.\n"
                                    "config.json gravado.\n\n"
                                    "Não consegui registrar automaticamente no Agendador do Windows:\n"
                                    + "\n".join(f"  •  {e}" for e in erros))
        else:
            messagebox.showinfo("Configuração concluída",
                                 f"{total_empresas} empresa(s) em {len(self._grupos_existentes())} grupo(s): {resumo_grupos}.\n\n"
                                 "config.json gravado e as tarefas foram registradas no Agendador\n"
                                 "do Windows automaticamente — cada grupo com sua própria agenda.\n"
                                 "Não precisa fazer mais nada.\n\n"
                                 f"A busca do histórico inicial ({resumo_periodo}) já começou, "
                                 "rodando em segundo plano — pode levar bastante tempo dependendo do "
                                 "volume de notas. Um aviso do Windows avisa quando terminar.")


def _despachar(subcomando: str) -> None:
    """Reencaminha para o main() de outro script deste projeto, dentro do
    MESMO processo — é o que permite empacotar tudo num .exe único: em vez
    de abrir um novo processo Python apontando pra um arquivo .py solto
    (que não existiria dentro do .exe), a gente só troca o sys.argv e
    chama a função diretamente.

    IMPORTANTE: repassa o código de retorno de main() para sys.exit() —
    sem isso, o processo sempre "parece" ter dado certo pra quem chamou
    (ex.: rotina.py), mesmo quando main() sinalizou falha (return 1)."""
    sys.argv = [sys.argv[0]] + sys.argv[2:]
    modulo = __import__(subcomando)
    codigo = modulo.main()
    sys.exit(codigo if codigo is not None else 0)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] in SUBCOMANDOS:
        _despachar(sys.argv[1])
    else:
        Assistente().mainloop()
